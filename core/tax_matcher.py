"""
税收分类匹配模块 - 根据 tax_list.xlsx 匹配物品的税收分类编码

加载 tax_list.xlsx（4207行），通过关键词模糊匹配，
为每个物品找到最合适的税收分类编码。
"""

import os
import re
import openpyxl

# 项目根目录
_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_DEFAULT_TAX_LIST = os.path.join(_PROJECT_ROOT, 'data', 'tax_list.xlsx')
_DEFAULT_TRAIN_DIR = os.path.join(_PROJECT_ROOT, 'data', 'train_data')


class TaxMatcher:
    """税收分类匹配器"""

    def __init__(self, tax_list_path: str = _DEFAULT_TAX_LIST,
                 train_dir: str = _DEFAULT_TRAIN_DIR):
        self.entries: list[dict] = []
        self.train_map: dict[str, str] = {}  # 物品名 → 税收编码
        self._load(tax_list_path)
        self._load_train(train_dir)

    def _load(self, path: str):
        """加载税收分类表"""
        if not os.path.exists(path):
            raise FileNotFoundError(f'税收分类表不存在: {path}')

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            # L列(index 11) = 合并编码, M列(12) = 名称, N列(13) = 分类名, O列(14) = 说明
            if len(row) < 13:
                continue

            code = row[11]  # L列 - 合并编码
            name = row[12]  # M列 - 合并名称

            if code is None or name is None:
                continue

            desc = row[14] if len(row) > 14 and row[14] else ''  # O列 - 说明
            category = row[13] if len(row) > 13 and row[13] else ''  # N列 - 分类名

            # 大数字编码（18-19位）在 Excel 中以 float 存储，str() 会变成科学计数法
            if isinstance(code, (int, float)):
                code = str(int(code))

            self.entries.append({
                'code': code,
                'name': str(name),
                'category': str(category),
                'desc': str(desc),
            })

        wb.close()

        # 排序后标记叶子节点（具体商品编码，非汇总分类）
        self.entries.sort(key=lambda e: e['code'])
        for i, entry in enumerate(self.entries):
            code = entry['code']
            is_leaf = True
            if i + 1 < len(self.entries):
                nxt = self.entries[i + 1]['code']
                if nxt.startswith(code) and nxt != code:
                    is_leaf = False
            entry['is_leaf'] = is_leaf

    def _load_train(self, train_dir: str):
        """加载训练数据（手动确认过的物品→编码映射），作为优先匹配"""
        if not os.path.isdir(train_dir):
            return
        import glob
        for fpath in sorted(glob.glob(os.path.join(train_dir, '*.xlsx'))):
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
                if len(row) < 2:
                    continue
                name = row[0]
                code = row[1]
                if name and code:
                    if isinstance(code, (int, float)):
                        code = str(int(code))
                    self.train_map[str(name).strip()] = str(code).strip()
            wb.close()

    def _extract_keywords(self, text: str) -> list[str]:
        """
        从物品名中提取关键词（中文子串）。

        "空气开关" → ["空气开关", "空气", "开关"]
        "平板灯" → ["平板灯", "平板", "板灯", "灯"]
        """
        # 提取连续中文
        chinese_parts = re.findall(r'[\u4e00-\u9fff]+', text)
        keywords = []
        for part in chinese_parts:
            keywords.append(part)
            # 2-gram 分词
            if len(part) >= 2:
                for i in range(len(part) - 1):
                    keywords.append(part[i:i+2])
            # 尾部单字（产品类型）
            if len(part) >= 1:
                keywords.append(part[-1])

        # 去重并按长度降序（优先匹配长词）
        seen = set()
        unique = []
        for kw in sorted(keywords, key=len, reverse=True):
            if kw not in seen:
                seen.add(kw)
                unique.append(kw)
        return unique

    def match(self, item_name: str) -> tuple[str, str]:
        """
        为物品匹配税收分类编码。

        匹配策略：
        0. 训练数据精确匹配 → 最高优先级（手动确认过的）
        1. 完整物品名在分类名称中出现 → 次高优先级
        2. 完整物品名在说明中出现 → 再次优先级
        3. 多个关键词在同一分类中匹配 → 加分
        4. 必须至少有一个关键词匹配分类名称（非说明）
        """
        # 优先匹配训练数据
        if item_name in self.train_map:
            train_code = self.train_map[item_name]
            for entry in self.entries:
                try:
                    if str(int(float(entry['code']))) == train_code:
                        return train_code, entry['name']
                except (ValueError, OverflowError):
                    if entry['code'] == train_code:
                        return train_code, entry['name']
            return train_code, ''

        keywords = self._extract_keywords(item_name)

        if not keywords:
            return '', ''

        # 物品名的尾部通常是产品类型（如"空气开关"的"开关"、"平板灯"的"灯"）
        # 提取尾部1-3字作为产品类型关键词，给予更高权重
        suffix_keywords = set()
        for n in [1, 2, 3]:
            if len(item_name) >= n:
                suffix_keywords.add(item_name[-n:])

        best_match = None
        best_score = -1

        for entry in self.entries:
            ename = entry['name']
            desc = entry['desc']
            code = entry['code']
            score = 0
            has_name_match = False
            matched_keywords = 0

            # 完整物品名匹配（最高优先级）
            if item_name in ename or ename in item_name:
                score += len(item_name) * 10
                has_name_match = True
                matched_keywords = len(keywords)

            if item_name in desc:
                score += len(item_name) * 8
                matched_keywords = max(matched_keywords, 1)

            has_desc_match = False

            # 逐个关键词匹配
            for kw in keywords:
                # 跳过单字（除非是尾部产品类型词）
                if len(kw) < 2 and kw not in suffix_keywords:
                    continue
                weight = len(kw) * 3
                # 产品类型关键词（尾部）加权：越靠后权重越高
                if kw in suffix_keywords:
                    if kw == item_name[-1:]:      # 最后1字（核心产品类型）
                        weight += 10
                    elif kw == item_name[-2:]:     # 最后2字
                        weight += 6
                    elif kw == item_name[-3:]:     # 最后3字
                        weight += 4
                if kw in ename:
                    score += weight
                    has_name_match = True
                    matched_keywords += 1
                elif kw in desc:
                    score += weight * 0.6
                    has_desc_match = True
                    matched_keywords += 1

            # 名称匹配优先；说明匹配次之但不大幅惩罚
            if not has_name_match and not has_desc_match:
                score = 0
            elif not has_name_match:
                score *= 0.6

            # 多关键词同时匹配加分（说明语义一致性高）
            if matched_keywords >= 2:
                score *= 1.5

            # 优先叶子节点（具体商品编码，非汇总分类）
            if score > 0:
                if entry.get('is_leaf', True):
                    score += 8          # 强偏好叶子节点
                else:
                    score -= 10         # 惩罚汇总/父级分类
                # 优先商品编码（1开头）而非服务编码
                if code.startswith('109'):      # 机电设备
                    score += 3
                elif code.startswith('108'):    # 非金属矿物
                    score += 2
                elif code.startswith('1'):      # 其他商品
                    score += 2
                # 短名称轻微优先
                score -= len(ename) * 0.05

            if score > best_score:
                best_score = score
                best_match = entry

        if best_match and best_score > 0:
            return best_match['code'], best_match['name']

        return '', ''

    def match_items(self, items: list) -> list:
        """
        为多个物品匹配税收分类。

        Args:
            items: parser.Item 列表

        Returns:
            填充了 tax_code 和 tax_name 的 items
        """
        for item in items:
            code, name = self.match(item.name)
            item.tax_code = code
            item.tax_name = name
        return items


if __name__ == '__main__':
    matcher = TaxMatcher()

    test_names = [
        '插座',
        '插座',
        '空气开关',
        '平板灯',
        '热水器',
        '时控开关',
        '电线',
    ]

    for name in test_names:
        code, category = matcher.match(name)
        print(f'{name:10s} → {code}  {category}')
