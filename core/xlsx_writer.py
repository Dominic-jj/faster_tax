"""
Excel 生成模块 - 基于 invoice_template.xlsx 模板生成电子发票

读取模板文件，在第4行开始填入物品数据，生成新的 xlsx 文件。
"""

import os
import shutil
import tempfile
import openpyxl

# 项目根目录
_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_DEFAULT_TEMPLATE = os.path.join(_PROJECT_ROOT, 'data', 'invoice_template.xlsx')

# 模板中的列映射（Row 3 表头对应的列字母）
COLUMN_MAP = {
    'name': 'A',       # 项目名称
    'tax_code': 'B',   # 商品和服务税收分类编码
    'spec': 'C',       # 规格型号
    'unit': 'D',       # 单位
    'quantity': 'E',   # 商品数量
    'unit_price': 'F', # 商品单价
    'amount': 'G',     # 金额
    'tax_rate': 'H',   # 税率
    'discount': 'I',   # 折扣金额
    'policy': 'J',     # 优惠政策类型
    'coal': 'K',       # 煤炭种类
}

DATA_START_ROW = 4  # 数据从第4行开始（前3行是说明和表头）
TAX_RATE = '0.01'


def generate(items: list, output_path: str,
             template_path: str = _DEFAULT_TEMPLATE) -> str:
    """
    生成电子发票 xlsx 文件。

    Args:
        items: parser.Item 列表（应已填充 tax_code）
        output_path: 输出文件路径
        template_path: 模板文件路径

    Returns:
        生成的文件路径
    """
    if not os.path.exists(template_path):
        raise FileNotFoundError(f'模板文件不存在: {template_path}')

    # 复制模板到输出路径
    shutil.copy2(template_path, output_path)

    # 打开并编辑
    wb = openpyxl.load_workbook(output_path)
    ws = wb['1-明细模板']

    for i, item in enumerate(items):
        row = DATA_START_ROW + i

        ws[f'A{row}'] = item.name
        # 编码为纯数字字符串时，设为文本格式防止 Excel 显示科学计数法
        cell_b = ws[f'B{row}']
        cell_b.value = str(item.tax_code) if item.tax_code else ''
        cell_b.number_format = '@'
        ws[f'C{row}'] = item.spec
        ws[f'D{row}'] = item.unit
        ws[f'E{row}'] = item.quantity
        ws[f'F{row}'] = item.unit_price
        ws[f'G{row}'] = item.amount
        ws[f'H{row}'] = TAX_RATE
        # I, J, K 留空

    wb.save(output_path)
    wb.close()

    return output_path


def generate_to_bytes(items: list,
                      template_path: str = _DEFAULT_TEMPLATE) -> bytes:
    """
    生成 xlsx 文件并返回字节数据（用于 Streamlit 下载）。

    Args:
        items: parser.Item 列表
        template_path: 模板文件路径

    Returns:
        xlsx 文件的字节数据
    """
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        temp_path = f.name

    try:
        generate(items, temp_path, template_path)
        with open(temp_path, 'rb') as f:
            data = f.read()
    finally:
        if os.path.exists(temp_path):
            os.unlink(temp_path)

    return data


if __name__ == '__main__':
    from parser import Item

    test_items = [
        Item(name='插座', spec='16A', unit='个', quantity=4, unit_price=15, amount=60),
        Item(name='插座', spec='公牛有线14座', unit='个', quantity=4, unit_price=38, amount=152),
        Item(name='空气开关', spec='2P 324', unit='个', quantity=3, unit_price=30, amount=90),
        Item(name='平板灯', spec='30X60', unit='个', quantity=2, unit_price=50, amount=100),
        Item(name='热水器', spec='', unit='个', quantity=1, unit_price=37, amount=37),
        Item(name='时控开关', spec='', unit='个', quantity=1, unit_price=55, amount=55),
    ]

    output = os.path.join(_PROJECT_ROOT, 'output', 'output_test.xlsx')
    os.makedirs(os.path.dirname(output), exist_ok=True)
    generate(test_items, output)
    print(f'已生成: {output}')
