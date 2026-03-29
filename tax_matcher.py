"""
税收分类匹配模块 - 根据 tax_list.xlsx 匹配物品的税收分类编码

加载 tax_list.xlsx（4207行），通过关键词模糊匹配，
为每个物品找到最合适的税收分类编码。
"""

import os
import re
import openpyxl


class TaxMatcher:
    """税收分类匹配器"""

    def __init__(self, tax_list_path: str = 'tax_list.xlsx'):
        self.entries: list[dict] = []
        self._load(tax_list_path)

    def _load(self, path: str):
        """加载税收分类表"""
        if not os.path.exists(path):
            raise FileNotFoundError(f'税收分类表不存在: {path}')

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            # L列(index 11) = 合并编码, M列(12) = 名称, N列(13) = 分类名, O列(14) = 说明
            if len(row) < 13:
                continue

            code = row[11]  # L列 - 合并编码
            name = row[12]  # M列 - 合并名称

            if code is None or name is None:
                continue

            desc = row[14] if len(row) > 14 and row[14] else ''  # O列 - 说明
            category = row[13] if len(row) > 13 and row[13] else ''  # N列 - 分类名

            self.entries.append({
                'code': str(code),
                'name': str(name),
                'category': str(category),
                'desc': str(desc),
            })

        wb.close()

    def _extract_keywords(self, text: str) -> list[str]:
        """
        从物品名中提取关键词（中文子串）。

        "空气开关" → ["空气开关", "空气", "开关"]
        "平板灯" → ["平板灯", "平板", "板灯", "灯"]
        """
        # 提取连续中文
        chinese_parts = re.findall(r'[\u4e00-\u9fff]+', text)
        keywords = []
        for part in chinese_parts:
            keywords.append(part)
            # 2-gram 分词
            if len(part) >= 2:
                for i in range(len(part) - 1):
                    keywords.append(part[i:i+2])
            # 尾部单字（产品类型）
            if len(part) >= 1:
                keywords.append(part[-1])

        # 去重并按长度降序（优先匹配长词）
        seen = set()
        unique = []
        for kw in sorted(keywords, key=len, reverse=True):
            if kw not in seen:
                seen.add(kw)
                unique.append(kw)
        return unique

    def match(self, item_name: str) -> tuple[str, str]:
        """
        为物品匹配税收分类编码。

        匹配策略：
        1. 完整物品名在分类名称中出现 → 最高优先级
        2. 完整物品名在说明中出现 → 次高优先级
        3. 多个关键词在同一分类中匹配 → 加分
        4. 必须至少有一个关键词匹配分类名称（非说明）
        """
        keywords = self._extract_keywords(item_name)

        if not keywords:
            return '', ''

        # 物品名的尾部通常是产品类型（如"空气开关"的"开关"、"平板灯"的"灯"）
        # 提取尾部1-3字作为产品类型关键词，给予更高权重
        suffix_keywords = set()
        for n in [1, 2, 3]:
            if len(item_name) >= n:
                suffix_keywords.add(item_name[-n:])

        best_match = None
        best_score = -1

        for entry in self.entries:
            ename = entry['name']
            desc = entry['desc']
            code = entry['code']
            score = 0
            has_name_match = False
            matched_keywords = 0

            # 完整物品名匹配（最高优先级）
            if item_name in ename or ename in item_name:
                score += len(item_name) * 10
                has_name_match = True
                matched_keywords = len(keywords)

            if item_name in desc:
                score += len(item_name) * 5
                matched_keywords = max(matched_keywords, 1)

            # 逐个关键词匹配
            for kw in keywords:
                # 跳过单字（除非是尾部产品类型词）
                if len(kw) < 2 and kw not in suffix_keywords:
                    continue
                weight = len(kw) * 3
                # 产品类型关键词（尾部）加权：越靠后权重越高
                if kw in suffix_keywords:
                    if kw == item_name[-1:]:      # 最后1字（核心产品类型）
                        weight += 10
                    elif kw == item_name[-2:]:     # 最后2字
                        weight += 6
                    elif kw == item_name[-3:]:     # 最后3字
                        weight += 4
                if kw in ename:
                    score += weight
                    has_name_match = True
                    matched_keywords += 1
                elif kw in desc:
                    score += weight * 0.3

            # 必须至少有一个关键词命中分类名称
            if not has_name_match and score <= len(item_name) * 5:
                # 只有 desc 匹配而没有 name 匹配，降低权重
                score *= 0.3

            # 多关键词同时匹配加分（说明语义一致性高）
            if matched_keywords >= 2:
                score *= 1.5

            # 优先叶子节点（轻微加分）
            if score > 0:
                # 优先商品编码（1开头）而非服务编码
                if code.startswith('109'):      # 机电设备
                    score += 3
                elif code.startswith('108'):    # 非金属矿物
                    score += 2
                elif code.startswith('1'):      # 其他商品
                    score += 2
                # 短名称优先（更通用的分类更可能是正确匹配）
                score -= len(ename) * 0.05

            if score > best_score:
                best_score = score
                best_match = entry

        if best_match and best_score > 0:
            return best_match['code'], best_match['name']

        return '', ''

    def match_items(self, items: list) -> list:
        """
        为多个物品匹配税收分类。

        Args:
            items: parser.Item 列表

        Returns:
            填充了 tax_code 和 tax_name 的 items
        """
        for item in items:
            code, name = self.match(item.name)
            item.tax_code = code
            item.tax_name = name
        return items


if __name__ == '__main__':
    matcher = TaxMatcher()

    test_names = [
        '折産',
        '插座',
        '空气开关',
        '平板灯',
        '热水器',
        '时控开关',
        '电线',
    ]

    for name in test_names:
        code, category = matcher.match(name)
        print(f'{name:10s} → {code}  {category}')
